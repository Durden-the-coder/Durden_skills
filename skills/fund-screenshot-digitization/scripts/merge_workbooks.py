"""Merge page-level fund transaction workbooks into one statistics sheet."""

from __future__ import annotations

import argparse
import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PAGE_RE = re.compile(r"第(\d+)页")
HEADERS = [
    "页码", "页内序号", "交易日期", "交易时刻", "产品名称", "基金代码",
    "业务类型", "申请数值", "申请单位", "确认数值", "确认单位",
    "关联账户", "状态", "来源文件",
]


def get_page(path: Path) -> int:
    match = PAGE_RE.search(path.name)
    if not match:
        raise ValueError(f"missing page number: {path.name}")
    return int(match.group(1))


def artifact_score(path: Path) -> int:
    if "精细复核版" in path.name:
        return 40
    if "数字化_复核版" in path.name:
        return 30
    if "数字化" in path.name:
        return 20
    return 10


def choose_artifacts(directory: Path) -> dict[int, Path]:
    selected: dict[int, tuple[int, Path]] = {}
    for path in directory.glob("*.xlsx"):
        if path.name.startswith("~$") or "合并统计输入" in path.name:
            continue
        if not path.name.startswith("交易记录_") or not PAGE_RE.search(path.name):
            continue
        page = get_page(path)
        candidate = (artifact_score(path), path)
        if page not in selected or candidate[0] > selected[page][0]:
            selected[page] = candidate
    return {page: pair[1] for page, pair in selected.items()}


def header_map(ws):
    values = next(ws.iter_rows(min_row=5, max_row=5, values_only=True), ())
    return {str(v).strip(): i for i, v in enumerate(values) if v not in (None, "")}


def value(values, mapping, name):
    index = mapping.get(name)
    return values[index] if index is not None and index < len(values) else None


def normalize_code(code):
    if code in (None, ""):
        return None
    if isinstance(code, float) and code.is_integer():
        code = int(code)
    if isinstance(code, int):
        return f"{code:06d}"
    text = str(code).strip()
    return text.zfill(6) if text.isdigit() else text


def normalize_date(value_):
    if isinstance(value_, datetime):
        return value_.date()
    if isinstance(value_, date):
        return value_
    return value_


def normalize_time(value_):
    if isinstance(value_, datetime):
        return value_.time()
    if isinstance(value_, time):
        return value_
    if isinstance(value_, str) and value_.strip():
        try:
            return datetime.strptime(value_.strip(), "%H:%M:%S").time()
        except ValueError:
            return value_.strip()
    return value_


def parse_display_number(text):
    if not isinstance(text, str):
        return text, None
    match = re.match(r"\s*([+-]?[\d,]+(?:\.\d+)?)\s*(.*)\s*$", text)
    if not match:
        return (None, None)
    return float(match.group(1).replace(",", "")), match.group(2) or None


def iter_rows(path: Path, page: int):
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_name = "数字字段" if "数字字段" in book.sheetnames else "交易记录"
        ws = book[sheet_name]
        mapping = header_map(ws)
        for values in ws.iter_rows(min_row=6, values_only=True):
            row_no = value(values, mapping, "序号")
            if row_no in (None, ""):
                continue
            row = [
                page,
                int(row_no) if isinstance(row_no, (int, float)) else row_no,
                normalize_date(value(values, mapping, "交易日期")),
                normalize_time(value(values, mapping, "交易时刻")),
                value(values, mapping, "产品名称"),
                normalize_code(value(values, mapping, "基金代码")),
                value(values, mapping, "业务类型"),
                value(values, mapping, "申请数值"),
                value(values, mapping, "申请单位"),
                value(values, mapping, "确认数值"),
                value(values, mapping, "确认单位"),
                value(values, mapping, "关联账户"),
                value(values, mapping, "状态"),
                path.name,
            ]
            if sheet_name == "交易记录":
                raw_time = value(values, mapping, "交易发起时间")
                if row[2] is None and isinstance(raw_time, str):
                    parts = raw_time.split(" ", 1)
                    row[2] = parts[0]
                    row[3] = parts[1] if len(parts) == 2 else None
                request, request_unit = parse_display_number(value(values, mapping, "申请金额"))
                confirm, confirm_unit = parse_display_number(value(values, mapping, "确认金额"))
                row[7], row[8] = request, request_unit
                row[9], row[10] = confirm, confirm_unit
            yield row
    finally:
        book.close()


def build_output(rows, output: Path):
    book = Workbook()
    ws = book.active
    ws.title = "交易明细"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    widths = [8, 10, 13, 12, 36, 12, 24, 14, 10, 14, 10, 22, 18, 44]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        row[2].number_format = "yyyy-mm-dd"
        row[3].number_format = "hh:mm:ss"
        row[5].number_format = "@"
        row[7].number_format = "0.00"
        row[9].number_format = "0.00"
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--expected-pages", type=int)
    args = parser.parse_args()
    selected = choose_artifacts(args.input_dir)
    if args.expected_pages:
        expected = set(range(1, args.expected_pages + 1))
        missing = sorted(expected - set(selected))
        if missing:
            raise SystemExit(f"missing pages: {missing}")
    rows = []
    for page in sorted(selected):
        rows.extend(iter_rows(selected[page], page))
    rows.sort(key=lambda item: (item[0], item[1]))
    build_output(rows, args.output)
    print(f"pages={len(selected)} rows={len(rows)} output={args.output}")


if __name__ == "__main__":
    main()
