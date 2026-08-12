"""Validate page artifacts or a merged transaction workbook."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl

PAGE_RE = re.compile(r"第(\d+)页")
MERGED_HEADERS = [
    "页码", "页内序号", "交易日期", "交易时刻", "产品名称", "基金代码",
    "业务类型", "申请数值", "申请单位", "确认数值", "确认单位",
    "关联账户", "状态", "来源文件",
]


def page_from_name(path: Path):
    match = PAGE_RE.search(path.name)
    return int(match.group(1)) if match else None


def scan_formulas(path: Path):
    book = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        formulas = []
        for ws in book.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append(f"{ws.title}!{cell.coordinate}")
        return formulas
    finally:
        book.close()


def validate_pages(directory: Path, expected_pages: int | None):
    files = {}
    for path in directory.glob("*.xlsx"):
        page = page_from_name(path)
        if page is None or "合并统计输入" in path.name:
            continue
        files.setdefault(page, []).append(path)
    if expected_pages:
        missing = sorted(set(range(1, expected_pages + 1)) - set(files))
        if missing:
            raise SystemExit(f"missing pages: {missing}")
    errors = []
    counts = {}
    for page, candidates in sorted(files.items()):
        path = max(candidates, key=lambda p: ("精细复核版" in p.name, "数字化" in p.name, p.name))
        book = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            if "数字字段" not in book.sheetnames:
                errors.append(f"page {page}: missing 数字字段")
                continue
            ws = book["数字字段"]
            header = list(next(ws.iter_rows(min_row=5, max_row=5, values_only=True), ()))
            if "序号" not in header or "基金代码" not in header:
                errors.append(f"page {page}: invalid header")
            count = sum(1 for row in ws.iter_rows(min_row=6, values_only=True) if row and row[0] not in (None, ""))
            counts[page] = count
        finally:
            book.close()
        formulas = scan_formulas(path)
        if formulas:
            errors.append(f"page {page}: formulas found {formulas[:5]}")
    if errors:
        raise SystemExit("\n".join(errors))
    print(f"page_files={len(files)} short_pages=" + ",".join(f"{p}:{n}" for p, n in counts.items() if n != 20))


def validate_merged(path: Path, expected_pages: int | None):
    book = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        if book.sheetnames != ["交易明细"]:
            raise SystemExit(f"expected one sheet named 交易明细, got {book.sheetnames}")
        ws = book["交易明细"]
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        if header != MERGED_HEADERS:
            raise SystemExit(f"invalid merged header: {header}")
        seen = set()
        pages = set()
        rows = 0
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or values[0] in (None, ""):
                continue
            rows += 1
            key = (values[0], values[1])
            if key in seen:
                raise SystemExit(f"duplicate page-row: {key}")
            seen.add(key)
            pages.add(values[0])
            code = values[5]
            if code not in (None, "") and (not isinstance(code, str) or len(code) != 6):
                raise SystemExit(f"invalid fund code at row {rows + 1}: {code!r}")
        if expected_pages:
            missing = sorted(set(range(1, expected_pages + 1)) - pages)
            if missing:
                raise SystemExit(f"merged file missing pages: {missing}")
        print(f"merged_rows={rows} pages={len(pages)} sheet=交易明细")
    finally:
        book.close()
    formulas = scan_formulas(path)
    if formulas:
        raise SystemExit(f"merged formulas found: {formulas[:5]}")


def main():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input-dir", type=Path)
    group.add_argument("--merged", type=Path)
    parser.add_argument("--expected-pages", type=int)
    args = parser.parse_args()
    if args.input_dir:
        validate_pages(args.input_dir, args.expected_pages)
    else:
        validate_merged(args.merged, args.expected_pages)


if __name__ == "__main__":
    main()
